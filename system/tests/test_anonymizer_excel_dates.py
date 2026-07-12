# -*- coding: utf-8 -*-
# SPDX-License-Identifier: MIT
"""Regressionstest: native Excel-Datumszellen (datetime/date-Objekte,
kein Text) wurden von _anonymize_excel bisher NIE geprueft, da der
Code nur `isinstance(cell.value, str)` behandelte. Ein echtes
Geburtsdatum, das Excel als datetime-Zelle speichert (z.B. in
Zeitnachweis-Tabellen), blieb dadurch unanonymisiert im Klartext
erhalten -- auch nach erfolgreicher Anonymisierung der Datei."""

import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl
import pytest

SYSTEM_ROOT = Path(__file__).parent.parent
if str(SYSTEM_ROOT) not in sys.path:
    sys.path.insert(0, str(SYSTEM_ROOT))

from hub._services.document.anonymizer_service import AnonymProfile, DocumentAnonymizer


def _make_profile():
    return AnonymProfile(
        client_id="K_TEST",
        tarnname="Laura Schuster",
        fake_geburtsdatum="07.04.2014",
        mappings={
            "names": {"Chris Darrell": "Laura Schuster"},
            "dates": {"02.03.2014": "07.04.2014"},
            "addresses": {},
            "misc": {},
        },
    )


def test_anonymize_excel_replaces_native_datetime_cell(tmp_path):
    path = tmp_path / "Zeitnachweis.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Klient"
    ws["B1"] = datetime(2014, 3, 2)
    wb.save(str(path))

    anonymizer = DocumentAnonymizer()
    ok, count = anonymizer._anonymize_excel(path, _make_profile())
    assert ok
    assert count >= 1

    wb2 = openpyxl.load_workbook(str(path))
    cell_value = wb2.active["B1"].value
    assert cell_value.strftime("%d.%m.%Y") == "07.04.2014", (
        f"Echtes Geburtsdatum in nativer datetime-Zelle wurde nicht ersetzt: {cell_value!r}"
    )


def test_anonymize_excel_replaces_native_date_cell(tmp_path):
    path = tmp_path / "Zeitnachweis_date.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = date(2014, 3, 2)
    wb.save(str(path))

    anonymizer = DocumentAnonymizer()
    ok, count = anonymizer._anonymize_excel(path, _make_profile())
    assert ok
    assert count >= 1

    wb2 = openpyxl.load_workbook(str(path))
    cell_value = wb2.active["A1"].value
    assert cell_value.strftime("%d.%m.%Y") == "07.04.2014"
